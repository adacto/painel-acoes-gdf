#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador do Painel de Ações Sociais — GDF na sua Porta
Lê a matriz ACOES_SOCIAIS (órgão × RA = total de atendimentos) e regrava
index_social.html a partir de template_social.html.

Fonte da matriz: aba ACOES_SOCIAIS da planilha (se existir); senão, o arquivo
local passado em MATRIX_FILE (modelo). Links dos relatórios: aba Entregues.
Autoria: Adacto Artur Dornas de Oliveira — SEGOV
"""
import io, json, os, re, sys, urllib.request

# PARTICIPAÇÃO ESPECIAL — órgãos que estiveram nas ações com entregas de outra
# natureza (não somadas ao total de atendimentos a pessoas). CEB foi para o ranking
# (decisão: quantidade de atendimentos). NOVACAP fica fora (é assunto do painel de obras).
# Só a SLU tem quadro especial clicável (detalhe por região).
SLU_RAS = [
    {"ra": "Itapoã", "ed": 1, "servicos": [
        {"s": "Coleta convencional", "v": "256,4 t"}, {"s": "Remoção de entulho", "v": "95 t"},
        {"s": "Varrição manual", "v": "293,58 km"}, {"s": "Pintura/frisagem de meio-fio", "v": "359,59 km"},
        {"s": "Catação", "v": "480 m²"}, {"s": "Coleta de animais mortos", "v": "35"},
        {"s": "Papa-lixo", "v": "2 instalados"}]},
    {"ra": "Paranoá", "ed": 2, "servicos": [
        {"s": "Coleta convencional", "v": "166,07 t"}, {"s": "Remoção de entulho", "v": "150 t"},
        {"s": "Varrição manual", "v": "295,78 km"}, {"s": "Pintura/frisagem de meio-fio", "v": "59,25 km"},
        {"s": "Catação", "v": "600 m²"}, {"s": "Coleta de animais mortos", "v": "20"},
        {"s": "Papa-lixo", "v": "42 operantes"}]},
    {"ra": "Riacho Fundo II", "ed": 3, "servicos": [
        {"s": "Coleta convencional", "v": "183,22 t"}, {"s": "Remoção de entulho", "v": "55 t"},
        {"s": "Varrição manual", "v": "177,45 km"}, {"s": "Pintura/frisagem de meio-fio", "v": "35,11 km"},
        {"s": "Catação", "v": "600 m²"}, {"s": "Coleta de animais mortos", "v": "8"},
        {"s": "Papa-lixo", "v": "9 operantes"}, {"s": "Construção do Papa-Entulho", "v": "50%"}]},
    {"ra": "Samambaia", "ed": 4, "servicos": [
        {"s": "Coleta convencional", "v": "898,82 t"}, {"s": "Remoção de entulho", "v": "350 t"},
        {"s": "Varrição manual", "v": "998,45 km"}, {"s": "Pintura/frisagem de meio-fio", "v": "94,25 km"},
        {"s": "Catação", "v": "600 m²"}, {"s": "Coleta de animais mortos", "v": "25"},
        {"s": "Papa-lixo", "v": "25 operantes"}]},
    {"ra": "26 de Setembro", "ed": 5, "servicos": [
        {"s": "Remoção de entulho", "v": "45 t"},
        {"s": "Demais serviços", "v": "leitura em andamento"}]},
    {"ra": "Ceilândia", "ed": 6, "servicos": [
        {"s": "GDF de cara nova", "v": "25% do planejado"},
        {"s": "Demais serviços", "v": "leitura em andamento"}]},
    {"ra": "Planaltina", "ed": 7, "servicos": [
        {"s": "Coleta convencional", "v": "666,1 t"}, {"s": "Remoção de entulho", "v": "703,53 t"},
        {"s": "Catação", "v": "600 m²"}, {"s": "Coleta de animais mortos", "v": "22"},
        {"s": "Papa-lixo", "v": "25 operantes"}, {"s": "Construção do Papa-Entulho", "v": "15%"},
        {"s": "Varrição / Pintura", "v": "total indisponível no arquivo"}]},
    {"ra": "Plano Piloto", "ed": 8, "servicos": [
        {"s": "Coleta convencional", "v": "1.158,93 t"}, {"s": "Remoção de entulho", "v": "105 t"},
        {"s": "Coleta seletiva", "v": "40,74 t"}, {"s": "Catação", "v": "600 m²"},
        {"s": "Coleta de animais mortos", "v": "25"}, {"s": "Papa-lixo", "v": "93 operantes"},
        {"s": "Varrição / Pintura", "v": "total indisponível no arquivo"}]},
    {"ra": "São Sebastião", "ed": 9, "servicos": [
        {"s": "Coleta convencional", "v": "331,06 t"}, {"s": "Remoção de entulho", "v": "90 t"},
        {"s": "Catação", "v": "480 m²"}, {"s": "Coleta de animais mortos", "v": "22"},
        {"s": "Papa-lixo", "v": "33 operantes"}, {"s": "Varrição / Pintura", "v": "total indisponível no arquivo"}]},
    {"ra": "Recanto das Emas", "ed": 10, "servicos": [
        {"s": "Coleta convencional", "v": "229,15 t"}, {"s": "Catação", "v": "720 m²"},
        {"s": "Papa-lixo", "v": "12 operantes"}, {"s": "Varrição / Pintura", "v": "total indisponível no arquivo"}]},
]

ESPECIAL = [
    {"orgao": "018 - SLU", "area": "Zeladoria e limpeza urbana",
     "destaque": "coleta, entulho, varrição, catação e mais", "slu": True,
     "nota": "Vários serviços por região (toneladas, m², km). Clique para ver o detalhe por região."},
    {"orgao": "013 - DFLEGAL", "area": "Fiscalização (DF Legal)", "destaque": "386 ações fiscais · 6 RAs"},
    {"orgao": "037 - CULTURA", "area": "Cultura — Mala do Livro", "destaque": "≈2.500 livros · 5 RAs"},
    {"orgao": "048 - SERINTER", "area": "Relações internacionais", "destaque": "19 embaixadas"},
    {"orgao": "026 - CGDF", "area": "Controladoria — Ouvidoria", "destaque": "34 manifestações"},
    {"orgao": "039 - SEAGRI", "area": "Agricultura", "destaque": "estradas rurais e roçagem (km/m²)"},
    {"orgao": "003 - SSP", "area": "Segurança pública", "destaque": "videomonitoramento · 28 veículos"},
    {"orgao": "011 - CAESB", "area": "Saneamento — água", "destaque": "unidade móvel e chamados"},
    {"orgao": "042 - POLO RURAL", "area": "Polo Rural", "destaque": "estradas rurais (km)"},
    {"orgao": "045 - SEPE", "area": "Projetos Especiais", "destaque": "Adote uma Praça"},
    {"orgao": "031 - SEMA", "area": "Meio ambiente", "destaque": "orientação e educação ambiental"},
    {"orgao": "041 - TERRACAP", "area": "Patrimonial", "destaque": "cercamento e cessão de uso"},
    {"orgao": "029 - Casa Militar", "area": "Apoio institucional", "destaque": "apoio à organização das ações"},
]

SHEET_ID = "1kYjCiFh8R9EOgnwBMEFR7YtZpO9MvUwDSXRuqH5Lvc4"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "template_social.html")
OUT = os.path.join(HERE, "index.html")
MATRIX_FILE = os.path.join(HERE, "ACOES_SOCIAIS_modelo.xlsx")  # usado se a aba não existir

_MIN={"de","da","do","das","dos","e"}; _ROM={"ii","iii","iv","v","vi","vii","viii","ix","x","xi"}
def bonito(nome):
    out=[]
    for i,p in enumerate(str(nome).strip().split()):
        low=p.lower()
        out.append(p.upper() if low in _ROM else (low if (low in _MIN and i>0) else (p if p.isdigit() else p[:1].upper()+p[1:].lower())))
    return re.sub(r"(Riacho Fundo)\s+2\b", r"\1 II", " ".join(out))

def num(v):
    if v in (None,""): return 0
    if isinstance(v,(int,float)): return int(round(v))
    s=re.sub(r"[^\d]","",str(v)); return int(s) if s else 0

def data_br(v):
    s=str(v or ""); m=re.match(r"(\d{4})-(\d{2})-(\d{2})",s)
    return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else s

def url_hyperlink(cell):
    v=cell.value
    if isinstance(v,str):
        m=re.search(r'HYPERLINK\("([^"]+)"',v)
        if m: return m.group(1)
    return cell.hyperlink.target if cell.hyperlink else None

def main():
    import openpyxl
    print("Baixando planilha…")
    data=urllib.request.urlopen(urllib.request.Request(XLSX_URL,headers={"User-Agent":"Mozilla/5.0"}),timeout=60).read()
    wb=openpyxl.load_workbook(io.BytesIO(data),data_only=True)
    wf=openpyxl.load_workbook(io.BytesIO(data),data_only=False)

    # data do retrato
    atualizado=""
    for r in wb["Resumo"].iter_rows(values_only=True):
        for c in r:
            m=re.search(r"Atualizado em\s+(\d{2}/\d{2}/\d{4} \d{2}:\d{2})",str(c or ""))
            if m: atualizado=m.group(1);break
        if atualizado:break

    # links dos relatórios: (cod, orgao) -> url; e órgãos participantes (distintos)
    link={}; participantes=set()
    ev,ef=wb["Entregues"],wf["Entregues"]
    for i in range(4,ev.max_row+1):
        cod=str(ev.cell(i,1).value or "").strip().upper()
        if not re.match(r"^P0\d\d$",cod): continue
        org=str(ev.cell(i,3).value or "").strip()
        link[(cod,org)]=url_hyperlink(ef.cell(i,7))
        if org: participantes.add(org)

    # matriz ACOES_SOCIAIS: da aba (se houver) ou do arquivo modelo
    if "ACOES_SOCIAIS" in wb.sheetnames:
        ws=wb["ACOES_SOCIAIS"]; print("Lendo aba ACOES_SOCIAIS da planilha.")
    else:
        ws=openpyxl.load_workbook(MATRIX_FILE,data_only=True)["ACOES_SOCIAIS"]
        print(f"Aba não encontrada — usando modelo local {os.path.basename(MATRIX_FILE)}.")

    hdr=[str(c.value or "").strip() for c in ws[1]]
    ra_cols=[]  # (colIdx0, cod, ra)
    for j,h in enumerate(hdr):
        m=re.match(r"(P0\d\d)\s+(.*)",h)
        if m: ra_cols.append((j,m.group(1),bonito(m.group(2))))

    orgaos_raw=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row or not row[0]: continue
        org=str(row[0]).strip()
        if org.upper().startswith("TOTAL"): continue
        itens=[]
        for j,cod,ra in ra_cols:
            v=num(row[j]) if j<len(row) else 0
            if v>0:
                itens.append({"cod":cod,"ra":ra,"ed":int(cod[1:]),"atend":v,
                              "url":link.get((cod,org))})
        if itens:
            orgaos_raw.append({"n":org,"total":sum(i["atend"] for i in itens),"itens":itens})

    orgaos_raw.sort(key=lambda o:-o["total"])

    # timeline: total por RA (soma dos órgãos)
    ras=[]
    for j,cod,ra in ra_cols:
        tot=sum(i["atend"] for o in orgaos_raw for i in o["itens"] if i["cod"]==cod)
        ras.append({"cod":cod,"ed":int(cod[1:]),"ra":ra,"atend":tot})

    kpis={"atendimentos":sum(o["total"] for o in orgaos_raw),
          "ras":len(ras),  # todas as RAs contempladas (com edição), inclusive as de atendimento 0
          "orgaos":len(orgaos_raw),
          "participantes":len(participantes)}
    snap={"atualizado":atualizado,"kpis":kpis,"ras":ras,"orgaos":orgaos_raw,
          "especial":ESPECIAL,"slu_ras":SLU_RAS}

    tpl=open(TEMPLATE,encoding="utf-8").read()
    open(OUT,"w",encoding="utf-8").write(tpl.replace("__SNAP_JSON__",json.dumps(snap,ensure_ascii=False,indent=1)))
    print(f"OK -> {OUT}")
    print(f"   total atendimentos: {kpis['atendimentos']:,} | RAs: {kpis['ras']} | órgãos: {kpis['orgaos']}".replace(",","."))
    for o in orgaos_raw[:5]: print(f"   {o['n']}: {o['total']:,}".replace(",","."))

if __name__=="__main__":
    main()
