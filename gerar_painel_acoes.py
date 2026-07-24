#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador do Painel de Prestação de Contas — GDF na sua Porta
===========================================================
Le a planilha de Monitoramento (abas Resumo, Painel, Entregues, CONFIG) —
as mesmas que o Apps Script mantem vivas a partir do Drive — e regrava o
index.html a partir de template.html.

Uso:  python3 gerar_painel_acoes.py     (gera index.html)
Requisito: openpyxl   (pip install openpyxl)
Autoria: Adacto Artur Dornas de Oliveira — SEGOV
"""
import io, json, re, os, sys, urllib.request

SHEET_ID = "1kYjCiFh8R9EOgnwBMEFR7YtZpO9MvUwDSXRuqH5Lvc4"
TOTAL_RAS_MIN = 37  # piso; o total real é lido da planilha (Resumo "X de N")
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "template.html")
OUT = os.path.join(HERE, "cobertura.html")

_MIN = {"de", "da", "do", "das", "dos", "e"}
_ROM = {"ii", "iii", "iv", "v", "vi", "vii", "viii", "ix", "x", "xi"}


def bonito(nome):
    out = []
    for i, p in enumerate(str(nome).strip().split()):
        low = p.lower()
        if low in _ROM: out.append(p.upper())
        elif low in _MIN and i > 0: out.append(low)
        elif p.isdigit(): out.append(p)
        else: out.append(p[:1].upper() + p[1:].lower())
    s = " ".join(out)
    s = re.sub(r"(Riacho Fundo)\s+2\b", r"\1 II", s)  # padrão da marca
    return s


def data_br(v):
    """datetime/str '2026-07-13 16:11:00' -> '13/07/2026'."""
    if v is None or v == "": return ""
    s = str(v)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m: return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    return m.group(0) if m else s


def url_hyperlink(cell):
    v = cell.value
    if isinstance(v, str):
        m = re.search(r'HYPERLINK\("([^"]+)"', v)
        if m: return m.group(1)
    if cell.hyperlink: return cell.hyperlink.target
    return None


def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("Instale openpyxl:  pip install openpyxl")

    print("Baixando planilha de monitoramento…")
    req = urllib.request.Request(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"})
    data = urllib.request.urlopen(req, timeout=60).read()
    wb_v = openpyxl.load_workbook(io.BytesIO(data), data_only=True)   # valores
    wb_f = openpyxl.load_workbook(io.BytesIO(data), data_only=False)  # formulas (links)

    # ---- CONFIG: nome oficial da RA por codigo ----
    cfg = {}
    ws = wb_v["CONFIG"]
    hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        cod = str(d.get("Codigo_P") or "").strip().upper()
        if re.match(r"^P0\d\d$", cod):
            cfg[cod] = bonito(str(d.get("Nome_Pasta") or "").strip())

    # ---- RESUMO: KPIs + entregues/pendentes/total por edicao ----
    ws = wb_v["Resumo"]
    rows = list(ws.iter_rows(values_only=True))
    kpi = {"edicoes": 0, "orgaos": 0, "entregas": 0, "pendencias": 0, "novidades": 0}
    total_ras = 0
    atualizado = ""
    por_ed = {}
    modo_tab = False
    for r in rows:
        cells = [("" if c is None else str(c).strip()) for c in r]
        j = " | ".join(cells)
        m = re.search(r"Atualizado em\s+(\d{2}/\d{2}/\d{4} \d{2}:\d{2})", j)
        if m: atualizado = m.group(1)
        rot = cells[0] if cells else ""
        val = cells[1] if len(cells) > 1 else ""
        if rot.startswith("Edições abertas"):
            mm = re.search(r"(\d+)", val); kpi["edicoes"] = int(mm.group(1)) if mm else 0
            mn = re.search(r"de\s+(\d+)", val); total_ras = int(mn.group(1)) if mn else 0
        elif rot.startswith("Órgãos monitorados"):
            kpi["orgaos"] = int(float(val)) if val else 0
        elif rot.startswith("Entregas válidas"):
            kpi["entregas"] = int(float(val)) if val else 0
        elif rot.startswith("Pendências"):
            kpi["pendencias"] = int(float(val)) if val else 0
        elif rot.startswith("Novidades"):
            kpi["novidades"] = int(float(val)) if val else 0
        if rot == "Edição" and val == "RA":
            modo_tab = True; continue
        if modo_tab and re.match(r"^P0\d\d$", rot):
            ent = int(float(cells[2])) if len(cells) > 2 and cells[2] else 0
            pen = int(float(cells[3])) if len(cells) > 3 and cells[3] else 0
            tot = int(float(cells[4])) if len(cells) > 4 and cells[4] else ent + pen
            por_ed[rot] = {"entregues": ent, "pendentes": pen, "total": tot,
                           "ra": bonito(cfg.get(rot, val))}

    # ---- ENTREGUES: relatorios por edicao (orgao, arquivo, data, link) ----
    wv, wf = wb_v["Entregues"], wb_f["Entregues"]
    ent_por_ed = {}
    ult_por_ed = {}
    participantes = set()
    for i in range(4, wv.max_row + 1):
        cod = wv.cell(i, 1).value
        cod = str(cod).strip().upper() if cod else ""
        if not re.match(r"^P0\d\d$", cod):
            continue
        orgao = str(wv.cell(i, 3).value or "").strip()
        arq = str(wv.cell(i, 4).value or "").strip()
        raw_data = wv.cell(i, 6).value
        dbr = data_br(raw_data)
        url = url_hyperlink(wf.cell(i, 7))
        ent_por_ed.setdefault(cod, []).append({"n": orgao, "arq": arq, "data": dbr, "url": url})
        participantes.add(re.sub(r"^\d+\s*-\s*", "", orgao).strip().upper())
        iso = str(raw_data or "")
        if iso and iso > ult_por_ed.get(cod, ""):
            ult_por_ed[cod] = iso

    # ---- monta as edicoes (ordena por codigo; ordena orgaos por numero) ----
    editions = []
    for cod in sorted(por_ed):
        info = por_ed[cod]
        orgs = ent_por_ed.get(cod, [])
        orgs.sort(key=lambda o: int(re.match(r"0*(\d+)", o["n"]).group(1)) if re.match(r"\d", o["n"]) else 999)
        editions.append({
            "cod": cod, "ed": int(cod[1:]), "ra": info["ra"],
            "entregues": info["entregues"], "pendentes": info["pendentes"], "total": info["total"],
            "ultima": data_br(ult_por_ed.get(cod, "")), "orgaos": orgs,
        })

    # ---- PAINEL: matriz órgão × edição (status) -> visão por órgão + pendentes ----
    wp = wb_v["Painel"]
    hdr_row = None
    for i in range(1, wp.max_row + 1):
        if str(wp.cell(i, 1).value or "").strip() == "Órgão":
            hdr_row = i
            break
    matrix = {}   # orgao -> {cod: status}
    ed_cols = []  # (coluna, cod)
    if hdr_row:
        for c in range(2, wp.max_column + 1):
            m = re.match(r"(P0\d\d)", str(wp.cell(hdr_row, c).value or "").strip())
            if m:
                ed_cols.append((c, m.group(1)))
        for i in range(hdr_row + 1, wp.max_row + 1):
            org = str(wp.cell(i, 1).value or "").strip()
            if not org or org.upper().startswith("TOTAL"):
                continue
            matrix[org] = {cod: str(wp.cell(i, c).value or "").strip().upper() for c, cod in ed_cols}

    ed_meta = {e["cod"]: {"ra": e["ra"], "ed": e["ed"]} for e in editions}
    det = {(e["cod"], o["n"]): o for e in editions for o in e["orgaos"]}

    def onum(o):
        m = re.match(r"0*(\d+)", o)
        return int(m.group(1)) if m else 999

    orgaos = []
    for org in sorted(matrix, key=onum):
        itens, ent, total = [], 0, 0
        for cod in sorted(ed_meta):
            st = matrix[org].get(cod, "SEM PASTA")
            if st == "SEM PASTA":
                continue
            total += 1
            d = det.get((cod, org), {})
            if st == "ENTREGUE":
                ent += 1
            itens.append({"cod": cod, "ra": ed_meta[cod]["ra"], "ed": ed_meta[cod]["ed"],
                          "status": st, "arq": d.get("arq", ""), "data": d.get("data", ""),
                          "url": d.get("url")})
        if total:
            orgaos.append({"n": org, "total": total, "entregues": ent, "itens": itens})

    for e in editions:
        e["pendentesList"] = sorted([o for o in matrix if matrix[o].get(e["cod"]) == "PENDENTE"], key=onum)

    kpi["participantes"] = len(participantes)
    snap = {"totalRAs": max(TOTAL_RAS_MIN, total_ras), "atualizado": atualizado,
            "kpis": kpi, "editions": editions, "orgaos": orgaos}
    snap_json = json.dumps(snap, ensure_ascii=False, indent=1)

    tpl = open(TEMPLATE, encoding="utf-8").read()
    open(OUT, "w", encoding="utf-8").write(tpl.replace("__SNAP_JSON__", snap_json))

    print(f"OK -> {OUT}")
    print(f"   Edições: {len(editions)} | entregas: {kpi['entregas']} | "
          f"participantes: {kpi['participantes']} | atualizado: {atualizado}")


if __name__ == "__main__":
    main()
